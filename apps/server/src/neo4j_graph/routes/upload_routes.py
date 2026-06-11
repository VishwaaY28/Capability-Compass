"""
Upload and document ingestion routes for Compass Master.

Handles PDF/document upload, extraction, CSV upload, and import to Neo4j graph database.
"""

import os
import logging
import tempfile
import csv
import io
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pathlib import Path

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/upload", tags=["Upload"])


# ---------------------------------------------------------------------------
# FIBO ontology + Compass ingestion log endpoints
# ---------------------------------------------------------------------------

@router.get("/ontology")
async def get_ontology_descriptor(include_concepts: bool = Query(True)):
    """
    Return the loaded FIBO ontology descriptor used by the Compass guardrail.

    Includes ontology metadata (IRI, label, concept count, threshold) and,
    by default, the full list of concepts (label / definition / parents).
    """
    try:
        from utils.ontology import get_ontology_service
        ontology = get_ontology_service()
        payload = {"ontology": ontology.metadata()}
        if include_concepts:
            payload["concepts"] = ontology.get_concepts()
        return payload
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to read ontology: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/ontology/sync")
async def sync_ontology_to_neo4j(
    replace_existing: bool = Query(True, description="Wipe existing :OntologyConcept nodes before sync"),
):
    """
    Project the FIBO ontology into Neo4j as ``(:OntologyConcept)-[:SUBCLASS_OF]->(:OntologyConcept)``.

    Useful for browsing the ontology alongside ingested capabilities.
    """
    try:
        from utils.ontology import get_ontology_service
        ontology = get_ontology_service()
        summary = ontology.sync_to_neo4j(replace_existing=replace_existing)
        return {"status": "success", "summary": summary}
    except Exception as e:
        logger.error(f"Ontology sync failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/ontology/reload")
async def reload_ontology():
    """Force a re-read of the FIBO RDF file from disk (e.g. after replacing it)."""
    try:
        from utils.ontology import get_ontology_service
        ontology = get_ontology_service(reload=True)
        return {"status": "success", "ontology": ontology.metadata()}
    except Exception as e:
        logger.error(f"Ontology reload failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/ontology/score-document")
async def score_document_against_ontology(
    file: UploadFile = File(...),
    top_k: Optional[int] = Query(None, ge=1, le=50),
    doc_threshold: Optional[float] = Query(None, ge=0.0, le=1.0),
    chunk_threshold: Optional[float] = Query(None, ge=0.0, le=1.0),
):
    """
    Score an uploaded document against the FIBO ontology *without* extracting.

    Returns the document-level relevance verdict plus the top concepts and
    their evidence chunks. Useful as a dry-run before paying for an LLM
    extraction call.
    """
    try:
        from utils.deepagent_extractor import load_document
        from utils.ontology import get_ontology_service

        allowed_extensions = [".pdf", ".docx", ".txt"]
        file_ext = Path(file.filename or "").suffix.lower()
        if file_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {file_ext}. Allowed: {', '.join(allowed_extensions)}",
            )

        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        try:
            chunks = load_document(tmp_path)
            ontology = get_ontology_service()
            relevance = ontology.score_document(
                chunks,
                top_k=top_k,
                doc_threshold=doc_threshold,
                chunk_threshold=chunk_threshold,
            )
            return {
                "status": "success",
                "filename": file.filename,
                "ontology": ontology.metadata(),
                "document_relevance": relevance.to_dict(),
                "would_extract": relevance.is_relevant,
            }
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Document scoring failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/ingestion-logs")
async def list_ingestion_logs(limit: int = Query(50, ge=1, le=1000)):
    """
    Return the most recent Compass ingestion runs.

    Each entry includes the source document, file hash, configuration, the
    ontology metadata that was applied, the per-process guardrail outcome
    (candidates / accepted / rejected with scores), capability name, status,
    duration and any error.
    """
    try:
        from utils.ingestion_logger import read_ingestion_logs
        return read_ingestion_logs(limit=limit)
    except Exception as e:
        logger.error(f"Failed to read ingestion logs: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/repair-uids")
async def repair_duplicate_uids():
    """
    One-shot maintenance: scan every node label that the CSV/LLM importer
    writes to and re-issue a globally unique ``uid`` to nodes that share
    one with another node of the same label.

    Background: an earlier version of the batch CSV importer assigned
    ``uid = idx + 1`` starting from 1 for every import, which collided
    with UIDs already in the graph. The downstream delete API matches
    by ``uid`` and runs ``DETACH DELETE`` — so deleting one capability
    would also wipe any other node of the same label that had the
    duplicated UID. This endpoint repairs the existing graph in place
    (newer imports already use ``max(uid)+1``).

    Returns a per-label report of how many UID collisions were fixed.
    """
    try:
        from neo4j_graph.services.query_execution_service import Neo4jQueryService

        labels = [
            "Vertical",
            "SubVertical",
            "Capability",
            "Process",
            "Subprocess",
            "DataEntity",
            "DataElements",
            "OrganizationUnit",
            "ApplicationCatalog",
        ]

        report = {}
        svc = Neo4jQueryService()
        try:
            for label in labels:
                # Find groups where two or more nodes share the same uid
                duplicate_query = f"""
                MATCH (n:`{label}`)
                WHERE n.uid IS NOT NULL
                WITH n.uid AS uid, collect(n) AS nodes
                WHERE size(nodes) > 1
                RETURN uid, [x IN nodes | id(x)] AS internal_ids
                """
                duplicate_groups = svc.execute_cypher(duplicate_query) or []

                if not duplicate_groups:
                    report[label] = {"collisions": 0, "reassigned": 0}
                    continue

                # Current max uid on this label, so re-issues never re-collide
                max_query = f"MATCH (n:`{label}`) RETURN coalesce(max(n.uid), 0) AS max_uid"
                max_rows = svc.execute_cypher(max_query) or []
                next_uid = (max_rows[0]["max_uid"] or 0) + 1

                reassigned = 0
                for group in duplicate_groups:
                    # Keep the first node's uid, re-issue UIDs for the rest
                    for internal_id in group["internal_ids"][1:]:
                        svc.execute_cypher(
                            f"""
                            MATCH (n:`{label}`) WHERE id(n) = $iid
                            SET n.uid = $new_uid
                            """,
                            {"iid": internal_id, "new_uid": next_uid},
                        )
                        next_uid += 1
                        reassigned += 1

                report[label] = {
                    "collisions": len(duplicate_groups),
                    "reassigned": reassigned,
                }
                logger.info(
                    f"[UID repair] {label}: fixed {len(duplicate_groups)} collision group(s), "
                    f"reassigned {reassigned} node(s)"
                )
        finally:
            svc.close()

        return {"status": "success", "report": report}

    except Exception as e:
        logger.error(f"UID repair failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/pdf")
async def upload_and_extract_pdf(
    file: UploadFile = File(...),
    vertical: Optional[str] = Query(None),
    subvertical: Optional[str] = Query(None),
    extraction_depth: str = Query("data_element")
):
    """
    Upload a PDF/document and extract capability model using LLM.
    
    Streams JSONL events back to the client for real-time progress updates.
    
    Query Parameters:
    - vertical: Optional vertical name override
    - subvertical: Optional subvertical name override
    - extraction_depth: Level to extract to (capability, process, subprocess, data_entity, data_element)
    
    Returns:
        Streaming JSONL response with extraction progress events
    """
    try:
        from utils.deepagent_extractor import extract_capability_model
        
        # Validate file type
        allowed_extensions = [".pdf", ".docx", ".txt"]
        file_ext = Path(file.filename or "").suffix.lower()
        if file_ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {file_ext}. Allowed: {', '.join(allowed_extensions)}"
            )
        
        # Save uploaded file to temporary location
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        logger.info(f"Uploaded file saved to: {tmp_path}")
        logger.info(f"Extraction parameters - vertical: {vertical}, subvertical: {subvertical}, depth: {extraction_depth}")
        
        # Stream extraction events back to client
        async def event_generator():
            try:
                async for event in extract_capability_model(
                    file_path=tmp_path,
                    vertical=vertical,
                    subvertical=subvertical,
                    extraction_depth=extraction_depth,
                    skip_embeddings=False  # Generate embeddings for knowledge chunks
                ):
                    # Convert event to JSONL format
                    import json
                    yield json.dumps(event) + "\n"
            finally:
                # Clean up temporary file
                try:
                    os.unlink(tmp_path)
                    logger.info(f"Cleaned up temporary file: {tmp_path}")
                except Exception as e:
                    logger.warning(f"Failed to clean up temp file: {e}")
        
        return StreamingResponse(
            event_generator(),
            media_type="application/x-ndjson"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload and extraction failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import-to-graph")
async def import_to_graph(payload: dict):
    """
    Import extracted capability model to Neo4j graph database.
    
    Request Body:
    {
        "model_data": {...extracted capability model...},
        "chunks_path": "path/to/chunks.json" (optional)
    }
    
    Returns:
        Import summary with counts of created entities
    """
    try:
        from neo4j_graph.services.capability_service import CapabilityService
        from neo4j_graph.services.process_service import ProcessService
        
        model_data = payload.get("model_data")
        chunks_path = payload.get("chunks_path")
        
        if not model_data:
            raise HTTPException(status_code=400, detail="model_data is required")
        
        # Validate model structure
        if not isinstance(model_data, dict):
            raise HTTPException(status_code=400, detail="model_data must be a JSON object")
        
        required_fields = ["name", "vertical"]
        for field in required_fields:
            if field not in model_data:
                raise HTTPException(status_code=400, detail=f"Missing required field: {field}")
        
        logger.info(f"Importing capability model: {model_data.get('name')}")
        
        # Import capability and processes to Neo4j
        summary = await _import_model_to_neo4j(model_data, chunks_path)
        
        # Refresh the LLM catalog so new capabilities are immediately searchable
        try:
            from utils.llmthinking import azure_openai_thinking_client
            azure_openai_thinking_client.refresh_catalog()
        except Exception as e:
            logger.warning(f"Catalog refresh after import failed (non-critical): {e}")
        
        return {
            "status": "success",
            "message": "Successfully imported to graph database",
            "summary": summary
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Import to graph failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


async def _import_model_to_neo4j(model_data: dict, chunks_path: Optional[str] = None) -> dict:
    """
    Internal function to import capability model to Neo4j.
    
    Args:
        model_data: Extracted capability model
        chunks_path: Optional path to document chunks JSON
        
    Returns:
        Dictionary with import statistics
    """
    from neo4j_graph.services.capability_service import CapabilityService
    from neo4j_graph.services.process_service import ProcessService
    from neo4j_graph.services.vertical_service import VerticalService
    
    stats = {
        "capabilities_created": 0,
        "processes_created": 0,
        "subprocesses_created": 0,
        "data_entities_created": 0,
        "data_elements_created": 0,
        "chunks_imported": 0,
        "ontology_links_created": 0,
    }

    # If the model carries guardrail metadata we make a best-effort to ensure
    # the FIBO ontology is also present in the graph so :ALIGNED_WITH edges
    # have a target node. This is a no-op if Neo4j or the ontology service
    # are unavailable and never fails the import.
    try:
        if model_data.get("ontology_guardrail", {}).get("applied"):
            from neo4j_graph.services.query_execution_service import Neo4jQueryService
            check_svc = Neo4jQueryService()
            try:
                rows = check_svc.execute_cypher(
                    "MATCH (n:OntologyConcept) RETURN count(n) AS n"
                )
                concept_count = rows[0]["n"] if rows else 0
            finally:
                check_svc.close()

            if concept_count == 0:
                from utils.ontology import get_ontology_service
                ontology = get_ontology_service()
                ontology.sync_to_neo4j(replace_existing=False)
                logger.info("[Import] Auto-synced FIBO ontology to Neo4j (was empty)")
    except Exception as e:
        logger.warning(f"[Import] FIBO ontology auto-sync skipped: {e}")
    
    try:
        # Step 1: Ensure vertical exists
        vertical_name = model_data.get("vertical")
        if not vertical_name:
            raise ValueError("Vertical name is required")
        
        # Get or create vertical
        verticals = VerticalService.get_all_verticals()
        vertical = next((v for v in verticals if v.get("name") == vertical_name), None)
        if not vertical:
            # Get next UID for vertical
            query = "MATCH (v:Vertical) RETURN max(v.uid) AS max_uid"
            from neo4j_graph.services.query_execution_service import Neo4jQueryService
            svc = Neo4jQueryService()
            try:
                results = svc.execute_cypher(query)
                max_uid = results[0]["max_uid"] if results and results[0]["max_uid"] else 0
                new_uid = max_uid + 1
            finally:
                svc.close()
            
            vertical = VerticalService.create_vertical(name=vertical_name, uid=new_uid)
            logger.info(f"Created new vertical: {vertical_name}")
        
        # Step 2: Ensure subvertical exists (if provided)
        subvertical_name = model_data.get("subvertical")
        subvertical = None
        if subvertical_name:
            subverticals = VerticalService.get_subverticals_by_vertical(vertical["uid"])
            subvertical = next((sv for sv in subverticals if sv.get("name") == subvertical_name), None)
            if not subvertical:
                # Get next UID for subvertical
                query = "MATCH (sv:SubVertical) RETURN max(sv.uid) AS max_uid"
                from neo4j_graph.services.query_execution_service import Neo4jQueryService
                svc = Neo4jQueryService()
                try:
                    results = svc.execute_cypher(query)
                    max_uid = results[0]["max_uid"] if results and results[0]["max_uid"] else 0
                    new_uid = max_uid + 1
                finally:
                    svc.close()
                
                subvertical = VerticalService.create_subvertical(
                    name=subvertical_name,
                    uid=new_uid,
                    vertical_id=vertical["uid"]  # Use uid instead of .id
                )
                logger.info(f"Created new subvertical: {subvertical_name}")
        
        # Step 3: Create capability
        # Get next UID for capability
        query = "MATCH (c:Capability) RETURN max(c.uid) AS max_uid"
        from neo4j_graph.services.query_execution_service import Neo4jQueryService
        svc = Neo4jQueryService()
        try:
            results = svc.execute_cypher(query)
            max_uid = results[0]["max_uid"] if results and results[0]["max_uid"] else 0
            new_uid = max_uid + 1
        finally:
            svc.close()
        
        capability = CapabilityService.create_capability(
            name=model_data.get("name"),
            description=model_data.get("description", ""),
            uid=new_uid,
            subvertical_id=subvertical.get("uid") if subvertical else None
        )
        stats["capabilities_created"] = 1
        logger.info(f"Created capability: {capability.get('name')} (UID: {capability.get('uid')})")
        
        # Step 4: Import processes and subprocesses
        processes = model_data.get("processes", [])
        for proc_idx, proc_data in enumerate(processes):
            # Get next UID for process
            query = "MATCH (p:Process) RETURN max(p.uid) AS max_uid"
            svc = Neo4jQueryService()
            try:
                results = svc.execute_cypher(query)
                max_uid = results[0]["max_uid"] if results and results[0]["max_uid"] else 0
                new_uid = max_uid + 1 + proc_idx
            finally:
                svc.close()
            
            # Prepare subprocesses with UIDs and data entities
            subprocesses_data = []
            subprocesses = proc_data.get("subprocesses", [])
            for sub_idx, subproc_data in enumerate(subprocesses):
                # Get next UID for subprocess
                query = "MATCH (sp:Subprocess) RETURN max(sp.uid) AS max_uid"
                svc = Neo4jQueryService()
                try:
                    results = svc.execute_cypher(query)
                    max_uid = results[0]["max_uid"] if results and results[0]["max_uid"] else 0
                    sub_uid = max_uid + 1 + sub_idx
                finally:
                    svc.close()
                
                # Prepare data entities with UIDs
                data_entities_data = []
                data_entities = subproc_data.get("data_entities", [])
                for de_idx, de_data in enumerate(data_entities):
                    # Get next UID for data entity
                    query = "MATCH (de:DataEntity) RETURN max(de.uid) AS max_uid"
                    svc = Neo4jQueryService()
                    try:
                        results = svc.execute_cypher(query)
                        max_uid = results[0]["max_uid"] if results and results[0]["max_uid"] else 0
                        de_uid = max_uid + 1 + de_idx
                    finally:
                        svc.close()
                    
                    # Prepare data elements with UIDs
                    data_elements_data = []
                    data_elements = de_data.get("data_elements", [])
                    for elem_idx, elem_data in enumerate(data_elements):
                        # Get next UID for data element
                        query = "MATCH (elem:DataElements) RETURN max(elem.uid) AS max_uid"
                        svc = Neo4jQueryService()
                        try:
                            results = svc.execute_cypher(query)
                            max_uid = results[0]["max_uid"] if results and results[0]["max_uid"] else 0
                            elem_uid = max_uid + 1 + elem_idx
                        finally:
                            svc.close()
                        
                        data_elements_data.append({
                            "uid": elem_uid,
                            "name": elem_data.get("data_element_name", elem_data.get("name", "")),
                            "description": elem_data.get("data_element_description", elem_data.get("description", ""))
                        })
                    
                    data_entities_data.append({
                        "uid": de_uid,
                        "name": de_data.get("data_entity_name", de_data.get("name", "")),
                        "description": de_data.get("data_entity_description", de_data.get("description", "")),
                        "data_elements": data_elements_data if data_elements_data else []
                    })
                
                subprocesses_data.append({
                    "uid": sub_uid,
                    "name": subproc_data.get("name"),
                    "description": subproc_data.get("description", ""),
                    "category": subproc_data.get("category"),
                    "data_entities": data_entities_data if data_entities_data else []
                })
            
            # Create process with subprocesses and data entities
            process = ProcessService.create_process(
                name=proc_data.get("name"),
                level=proc_data.get("level", "core"),
                description=proc_data.get("description", ""),
                uid=new_uid,
                capability_id=capability.get("uid"),
                category=proc_data.get("category"),
                subprocesses=subprocesses_data if subprocesses_data else None
            )
            stats["processes_created"] += 1
            stats["subprocesses_created"] += len(subprocesses_data)

            # If this process was aligned with a FIBO concept by the
            # ontology guardrail, materialise the alignment in the graph.
            alignment = proc_data.get("ontology_alignment") or {}
            concept_iri = alignment.get("concept_iri")
            if concept_iri:
                try:
                    from neo4j_graph.services.query_execution_service import Neo4jQueryService
                    align_svc = Neo4jQueryService()
                    try:
                        align_svc.execute_cypher(
                            """
                            MATCH (p:Process {uid: $proc_uid})
                            MERGE (oc:OntologyConcept {iri: $iri})
                              ON CREATE SET oc.label = $label,
                                            oc.source = $source,
                                            oc.created_from = 'guardrail_import'
                            MERGE (p)-[r:ALIGNED_WITH]->(oc)
                            SET r.score = $score,
                                r.score_breakdown = $score_breakdown,
                                r.threshold = $threshold,
                                r.source = $source,
                                r.aligned_at = datetime()
                            SET p.ontology_concept_iri = $iri,
                                p.ontology_concept_label = $label,
                                p.ontology_score = $score
                            """,
                            {
                                "proc_uid": new_uid,
                                "iri": concept_iri,
                                "label": alignment.get("concept_label") or "",
                                "score": float(alignment.get("score") or 0.0),
                                "score_breakdown": str(alignment.get("score_breakdown") or {}),
                                "threshold": float(alignment.get("threshold") or 0.0),
                                "source": alignment.get("source") or "FIBO",
                            },
                        )
                        stats["ontology_links_created"] += 1
                    finally:
                        align_svc.close()
                except Exception as link_exc:
                    logger.warning(
                        f"[Import] Failed to link Process {new_uid} to ontology concept "
                        f"{concept_iri}: {link_exc}"
                    )
            
            # Count data entities and elements created
            for sp in subprocesses_data:
                for de in sp.get("data_entities", []):
                    stats["data_entities_created"] = stats.get("data_entities_created", 0) + 1
                    stats["data_elements_created"] = stats.get("data_elements_created", 0) + len(de.get("data_elements", []))
            
            logger.info(f"Created process: {process.get('name')} (UID: {process.get('uid')}) with {len(subprocesses_data)} subprocesses")
        
        # Step 5: Import document chunks as Chunk nodes (for RAG/knowledge retrieval)
        if chunks_path and os.path.exists(chunks_path):
            try:
                import json
                with open(chunks_path, 'r', encoding='cp1252') as f:
                    chunks_data = json.load(f)
                    chunks = chunks_data.get("chunks", [])
                    
                    if chunks:
                        logger.info(f"Importing {len(chunks)} knowledge chunks to Neo4j...")
                        
                        # Import chunks to Neo4j as Chunk nodes
                        from neo4j_graph.services.query_execution_service import Neo4jQueryService
                        svc = Neo4jQueryService()
                        try:
                            # Create Chunk nodes and link them to the capability
                            for idx, chunk in enumerate(chunks):
                                chunk_text = chunk.get("text", "")
                                chunk_metadata = chunk.get("metadata", {})
                                chunk_embedding = chunk.get("embedding", [])
                                
                                if not chunk_text:
                                    continue
                                
                                # Validate embedding is a list of numbers
                                if chunk_embedding and isinstance(chunk_embedding, list) and len(chunk_embedding) > 0:
                                    # Ensure all elements are numbers
                                    try:
                                        embedding_valid = all(isinstance(x, (int, float)) for x in chunk_embedding)
                                        if not embedding_valid:
                                            logger.warning(f"Chunk {idx} has invalid embedding format, skipping embedding")
                                            chunk_embedding = []
                                    except Exception:
                                        chunk_embedding = []
                                else:
                                    chunk_embedding = []
                                
                                # Create Chunk node with text, metadata, and embedding
                                # Note: Neo4j stores embeddings as a list property
                                create_chunk_query = """
                                MATCH (c:Capability {uid: $capability_uid})
                                CREATE (chunk:Chunk {
                                    uid: $chunk_uid,
                                    text: $text,
                                    page: $page,
                                    source: $source,
                                    embedding: $embedding
                                })
                                CREATE (c)-[:HAS_CHUNK]->(chunk)
                                RETURN chunk.uid as chunk_uid
                                """
                                
                                params = {
                                    "capability_uid": capability.get("uid"),
                                    "chunk_uid": idx + 1,
                                    "text": chunk_text,
                                    "page": chunk_metadata.get("page", 0),
                                    "source": chunks_data.get("metadata", {}).get("source_file", "unknown"),
                                    "embedding": chunk_embedding  # Pass as list directly
                                }
                                
                                result = svc.execute_cypher(create_chunk_query, params)
                                
                                # Log first chunk for verification
                                if idx == 0:
                                    logger.info(f"First chunk created - UID: {idx + 1}, embedding length: {len(chunk_embedding)}")
                            
                            stats["chunks_imported"] = len(chunks)
                            logger.info(f"Successfully imported {len(chunks)} knowledge chunks to Neo4j with embeddings")
                            
                        finally:
                            svc.close()
                            
            except Exception as e:
                logger.warning(f"Failed to import chunks: {e}", exc_info=True)
                # Don't fail the entire import if chunks fail
        
        return stats
        
    except Exception as e:
        logger.error(f"Failed to import model to Neo4j: {e}", exc_info=True)
        raise



TABULAR_EXTENSIONS = {".csv", ".xlsx", ".xls"}


def _xlsx_to_csv_text(xlsx_bytes: bytes) -> str:
    """
    Convert XLSX/XLS bytes to CSV text using openpyxl.

    Any leading rows that are completely empty are skipped so the first
    *non-empty* row becomes the CSV header. Cells containing ``None`` are
    emitted as empty strings.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.active

    output = io.StringIO()
    writer = csv.writer(output)
    seen_non_empty = False
    for row in ws.iter_rows(values_only=True):
        cells = ["" if cell is None else str(cell) for cell in row]
        if not seen_non_empty:
            if not any(c.strip() for c in cells):
                continue
            seen_non_empty = True
        writer.writerow(cells)

    return output.getvalue()


def _read_tabular_file_as_csv(filename: str, content: bytes) -> str:
    """
    Normalise a CSV/XLSX upload to CSV text so the rest of the pipeline can
    treat it uniformly. Raises ``HTTPException`` for unsupported types.
    """
    file_ext = Path(filename or "").suffix.lower()
    if file_ext not in TABULAR_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Unsupported tabular file type: {file_ext}. "
                f"Allowed: {', '.join(sorted(TABULAR_EXTENSIONS))}"
            ),
        )

    if file_ext in {".xlsx", ".xls"}:
        try:
            return _xlsx_to_csv_text(content)
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Failed to parse spreadsheet: {e}",
            )

    # Try the common encodings; `utf-8-sig` first so a BOM at the very top
    # of the file is stripped automatically.
    for encoding in ("utf-8-sig", "cp1252", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


# Map of accepted header *aliases* → the canonical column name that
# `CSVImportService` expects. Each alias is matched after normalization
# (lower-cased, whitespace-collapsed, hyphens treated like spaces) so
# minor casing / punctuation / pluralisation differences in the source
# spreadsheet do not require the user to rename headers.
HEADER_ALIASES: Dict[str, List[str]] = {
    "Vertical": ["vertical","VERTICAL"],
    "Sub Vertical": ["sub vertical", "subvertical", "SUBVERTICAL","SUB-VERTICAL","sub-vertical","Sub-Vertical"],
    "Capability Name": ["capability name", "capability", "CAPABILITY NAME", "CAPABILITY"],
    "Process": ["process", "process name", "PROCESS", "PROCESS NAME"],
    "Process Description": ["process description", "PROCESS DESCRIPTION", "PROCESS-DESCRIPTION","process-description"],
    "Sub Process": [
        "sub process",
        "subprocess",
        "sub process name",
        "subprocess name",
        "sub-process"
    ],
    "Sub-Process Description": [
        "sub process description",
        "subprocess description",
        "sub-process description",
    ],
    "Data Entity": ["data entity", "dataentity"],
    "Data Entity Description": ["data entity description"],
    "Data Element": ["data element", "dataelement"],
    "Data Element Description": ["data element description"],
    "Organization Units": [
        "organization units",
        "organisation units",
        "org units",
        "organizational units",
        "organisational units",
        "organization unit",
        "organisation unit",
        "org unit",
    ],
    "Applications": [
        "applications",
        "application",
        "apps",
        "app",
        "supporting applications",
        "supporting application",
    ],
}


def _normalize_csv_headers(csv_text: str) -> str:
    """
    Rewrite the first non-empty row of ``csv_text`` so that any header
    whose normalised form matches one of the canonical
    ``CSVImportService.EXPECTED_COLUMNS`` (directly or via
    ``HEADER_ALIASES``) is replaced by the canonical casing/spacing that
    the validator and parser expect.

    Normalisation: trim, lower-case, collapse runs of whitespace, treat
    hyphens like spaces (so ``"Sub-Process"`` and ``"Sub Process"``
    collide), so headers like ``"Application"`` are accepted as
    ``"Applications"``.

    Also strips a UTF-8 BOM from the very first character.

    Unknown headers are passed through unchanged so partial files still
    surface a descriptive "Missing columns" error.
    """
    import re
    from neo4j_graph.services.csv_import_service import CSVImportService

    if not csv_text:
        return csv_text

    if csv_text.startswith("\ufeff"):
        csv_text = csv_text[1:]

    reader = csv.reader(io.StringIO(csv_text))
    try:
        first_row = next(reader)
    except StopIteration:
        return csv_text

    def _norm(s: str) -> str:
        s = (s or "").replace("-", " ")
        return re.sub(r"\s+", " ", s).strip().lower()

    # Build the alias → canonical lookup: every expected column maps to
    # itself, then every entry in HEADER_ALIASES (re-normalised so the
    # author of the alias map doesn't have to be exact).
    alias_to_canonical: Dict[str, str] = {
        _norm(col): col for col in CSVImportService.EXPECTED_COLUMNS
    }
    for canonical_name, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_to_canonical[_norm(alias)] = canonical_name

    rewritten_headers = [alias_to_canonical.get(_norm(h), h) for h in first_row]

    body_io = io.StringIO()
    writer = csv.writer(body_io, lineterminator="\n")
    writer.writerow(rewritten_headers)
    for remaining in reader:
        writer.writerow(remaining)
    return body_io.getvalue()


def _build_capabilities_from_csv_text(csv_text: str) -> List[Dict[str, Any]]:
    """
    Group already-validated CSV rows into ``ExtractedCapabilityModel``-shaped
    dictionaries — one per distinct ``Capability Name`` — so the UI can
    render them the same way it renders LLM-extracted models.

    No Neo4j writes happen here; this is a pure transformation.
    """
    from neo4j_graph.services.csv_import_service import CSVImportService

    rows = CSVImportService.parse_csv_rows(csv_text)

    capabilities_map: Dict[str, Dict[str, Any]] = {}
    processes_map: Dict[tuple, Dict[str, Any]] = {}
    subprocesses_map: Dict[tuple, Dict[str, Any]] = {}
    data_entities_map: Dict[tuple, Dict[str, Any]] = {}
    seen_elements: set = set()

    for row in rows:
        cap_name = row["capability_name"]
        if not cap_name:
            continue

        cap = capabilities_map.get(cap_name)
        if cap is None:
            cap = {
                "name": cap_name,
                "description": "",
                "vertical": row["vertical"],
                "subvertical": row["subvertical"],
                "processes": [],
            }
            capabilities_map[cap_name] = cap
        else:
            if not cap["vertical"] and row["vertical"]:
                cap["vertical"] = row["vertical"]
            if not cap["subvertical"] and row["subvertical"]:
                cap["subvertical"] = row["subvertical"]

        proc_name = row["process_name"]
        if not proc_name:
            continue

        pkey = (cap_name, proc_name)
        proc = processes_map.get(pkey)
        if proc is None:
            proc = {
                "name": proc_name,
                "level": "core",
                "description": row["process_desc"],
                "category": None,
                "subprocesses": [],
            }
            processes_map[pkey] = proc
            cap["processes"].append(proc)
        elif not proc["description"] and row["process_desc"]:
            proc["description"] = row["process_desc"]

        sub_name = row["subprocess_name"]
        if not sub_name:
            continue

        skey = (cap_name, proc_name, sub_name)
        sub = subprocesses_map.get(skey)
        if sub is None:
            sub = {
                "name": sub_name,
                "description": row["subprocess_desc"],
                "category": None,
                "data_entities": [],
            }
            subprocesses_map[skey] = sub
            proc["subprocesses"].append(sub)
        elif not sub["description"] and row["subprocess_desc"]:
            sub["description"] = row["subprocess_desc"]

        de_name = row["data_entity_name"]
        if not de_name:
            continue

        dekey = (cap_name, proc_name, sub_name, de_name)
        de = data_entities_map.get(dekey)
        if de is None:
            de = {
                "data_entity_name": de_name,
                "data_entity_description": row["data_entity_desc"],
                "data_elements": [],
            }
            data_entities_map[dekey] = de
            sub["data_entities"].append(de)
        elif not de["data_entity_description"] and row["data_entity_desc"]:
            de["data_entity_description"] = row["data_entity_desc"]

        elem_name = row["data_element_name"]
        if not elem_name:
            continue

        elemkey = (cap_name, proc_name, sub_name, de_name, elem_name)
        if elemkey in seen_elements:
            continue
        seen_elements.add(elemkey)
        de["data_elements"].append({
            "data_element_name": elem_name,
            "data_element_description": row["data_element_desc"],
        })

    return list(capabilities_map.values())


def _aggregate_capabilities_into_single_model(
    capabilities: List[Dict[str, Any]],
    filename: str,
) -> Dict[str, Any]:
    """
    Collapse one-or-many capabilities parsed from a CSV/XLSX into a single
    ``ExtractedCapabilityModel`` shape so the existing single-capability
    ingestion modal can render the result without any special casing.

    - If the file contains exactly one capability, it is returned as-is.
    - If it contains multiple, they are merged: the file stem becomes the
      model name, vertical/sub-vertical fall back to "Multiple" when the
      source rows disagree, and every process is appended under a flat
      ``processes`` list in the order it appeared.
    """
    if not capabilities:
        raise HTTPException(
            status_code=400,
            detail="No capability rows found in the file (every row was empty or missing a Capability Name).",
        )

    if len(capabilities) == 1:
        only = dict(capabilities[0])
        if not only.get("description"):
            only["description"] = f"Imported from {filename}"
        return only

    verticals = {c.get("vertical") for c in capabilities if c.get("vertical")}
    subverticals = {c.get("subvertical") for c in capabilities if c.get("subvertical")}

    merged_vertical = verticals.pop() if len(verticals) == 1 else "Multiple"
    merged_subvertical = subverticals.pop() if len(subverticals) == 1 else "Multiple"

    aggregated_processes: List[Dict[str, Any]] = []
    for cap in capabilities:
        for proc in cap.get("processes", []):
            aggregated_processes.append(proc)

    stem = Path(filename).stem or filename
    cap_names = ", ".join(c["name"] for c in capabilities)

    return {
        "name": stem,
        "description": f"Imported from {filename} ({len(capabilities)} capabilities: {cap_names})",
        "vertical": merged_vertical,
        "subvertical": merged_subvertical,
        "processes": aggregated_processes,
    }


@router.post("/tabular-preview")
async def preview_tabular_file(file: UploadFile = File(...)):
    """
    Parse a CSV/XLSX capability mapping file and return the hierarchical
    structure WITHOUT running the LLM extractor, the FIBO ontology guardrail,
    or any Neo4j writes.

    The response shape mirrors the LLM-extracted ``ExtractedCapabilityModel``
    so the existing ingestion modal can render it unchanged. If the file
    contains multiple capabilities they are merged into one model whose
    ``processes`` list is the concatenation of every capability's processes.
    """
    try:
        from neo4j_graph.services.csv_import_service import CSVImportService

        if not file.filename:
            raise HTTPException(status_code=400, detail="Filename is required")

        content = await file.read()
        csv_text = _read_tabular_file_as_csv(file.filename, content)
        csv_text = _normalize_csv_headers(csv_text)

        is_valid, error_msg = CSVImportService.validate_csv_structure(csv_text)
        if not is_valid:
            # Surface the actual header row we saw so the user can spot the
            # mismatch (extra spaces, different casing, wrong sheet, etc.).
            try:
                first_line = csv_text.splitlines()[0] if csv_text else ""
                observed_headers = next(csv.reader(io.StringIO(first_line)))
            except Exception:
                observed_headers = []
            logger.warning(
                f"[Tabular preview] '{file.filename}' rejected: {error_msg}. "
                f"Observed headers: {observed_headers}"
            )
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Invalid tabular file structure: {error_msg}. "
                    f"Found columns: {observed_headers}"
                ),
            )

        capabilities = _build_capabilities_from_csv_text(csv_text)
        model = _aggregate_capabilities_into_single_model(capabilities, file.filename)

        logger.info(
            f"Tabular preview parsed '{file.filename}': "
            f"{len(capabilities)} capability/ies, {len(model['processes'])} process(es)"
        )

        return {
            "status": "success",
            "filename": file.filename,
            "model": model,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Tabular preview failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/csv")
async def upload_csv(
    file: UploadFile = File(...),
    clear_existing: bool = Query(False, description="Clear existing data before import"),
):
    """
    Upload a CSV or XLSX file and import the capability model to Neo4j.
    
    Expected columns (header row):
    - Vertical
    - Sub Vertical
    - Capability Name
    - Process
    - Process Description
    - Sub Process
    - Sub-Process Description
    - Data Entity
    - Data Entity Description
    - Data Element
    - Data Element Description
    - Organization Units (comma-separated)
    - Applications (comma-separated)
    
    Query Parameters:
    - clear_existing: If true, clears all existing data before import
    
    Returns:
        Import summary with counts of created entities
    """
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="Filename is required")

        content = await file.read()
        csv_text = _read_tabular_file_as_csv(file.filename, content)
        csv_text = _normalize_csv_headers(csv_text)

        logger.info(f"Processing tabular file: {file.filename}")

        # Always use the row-by-row CSVImportService path: it calls
        # get_max_uids() so newly created nodes get globally-unique UIDs
        # (the batch path used idx+1 which collided with existing UIDs
        # and caused delete-by-uid to wipe unrelated nodes).
        summary = await _import_csv_to_neo4j(csv_text, clear_existing)
        
        # Refresh the LLM catalog so newly imported capabilities are immediately searchable
        try:
            from utils.llmthinking import azure_openai_thinking_client
            azure_openai_thinking_client.refresh_catalog()
        except Exception as e:
            logger.warning(f"Catalog refresh after CSV import failed (non-critical): {e}")
        
        return {
            "status": "success",
            "message": "Successfully imported tabular file to graph database",
            "summary": summary
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"CSV upload failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


async def _import_csv_to_neo4j(csv_text: str, clear_existing: bool = False) -> dict:
    """
    Parse CSV and import to Neo4j graph database.
    
    Args:
        csv_text: CSV file content as string
        clear_existing: Whether to clear existing data first
        
    Returns:
        Dictionary with import statistics
    """
    from neo4j_graph.services.csv_import_service import CSVImportService
    
    stats = {
        "verticals_created": 0,
        "subverticals_created": 0,
        "capabilities_created": 0,
        "processes_created": 0,
        "subprocesses_created": 0,
        "data_entities_created": 0,
        "data_elements_created": 0,
        "organization_units_created": 0,
        "applications_created": 0,
        "rows_processed": 0
    }
    
    try:
        # Validate CSV structure
        is_valid, error_msg = CSVImportService.validate_csv_structure(csv_text)
        if not is_valid:
            raise ValueError(error_msg)
        
        # Clear existing data if requested (no re-seeding)
        if clear_existing:
            CSVImportService.clear_all_data()
            logger.info("Cleared all existing data - will import from CSV only")
        
        # Get starting UIDs
        uid_counters = CSVImportService.get_max_uids()
        
        # Track created entities to avoid duplicates
        verticals_cache = {}  # name -> uid
        subverticals_cache = {}  # (vertical_uid, name) -> uid
        capabilities_cache = {}  # name -> uid
        processes_cache = {}  # (capability_uid, name) -> uid
        subprocesses_cache = {}  # (process_uid, name) -> uid
        data_entities_cache = {}  # (subprocess_uid, name) -> uid
        data_elements_cache = {}  # (data_entity_uid, name) -> uid
        org_units_cache = {}  # name -> uid
        applications_cache = {}  # name -> uid
        
        # Parse CSV rows
        rows = CSVImportService.parse_csv_rows(csv_text)
        logger.info(f"Parsed {len(rows)} valid rows from CSV")
        
        # Process each row
        for idx, row in enumerate(rows):
            if idx % 50 == 0:
                logger.info(f"Processing row {idx + 1}/{len(rows)}...")
            
            stats["rows_processed"] += 1
            
            vertical_name = row["vertical"]
            subvertical_name = row["subvertical"]
            capability_name = row["capability_name"]
            process_name = row["process_name"]
            process_desc = row["process_desc"]
            subprocess_name = row["subprocess_name"]
            subprocess_desc = row["subprocess_desc"]
            data_entity_name = row["data_entity_name"]
            data_entity_desc = row["data_entity_desc"]
            data_element_name = row["data_element_name"]
            data_element_desc = row["data_element_desc"]
            org_units_str = row["org_units"]
            applications_str = row["applications"]
            
            # 1. Create or get Vertical
            vertical_uid = None
            if vertical_name:
                if vertical_name not in verticals_cache:
                    CSVImportService.create_or_get_vertical(
                        vertical_name, uid_counters, verticals_cache
                    )
                    stats["verticals_created"] += 1
                vertical_uid = verticals_cache[vertical_name]
            
            # 2. Create or get SubVertical
            subvertical_uid = None
            if subvertical_name and vertical_uid:
                subvertical_key = (vertical_uid, subvertical_name)
                if subvertical_key not in subverticals_cache:
                    CSVImportService.create_or_get_subvertical(
                        subvertical_name, vertical_uid, uid_counters, subverticals_cache
                    )
                    stats["subverticals_created"] += 1
                subvertical_uid = subverticals_cache[subvertical_key]
            
            # 3. Create or get Capability (linked to SubVertical)
            if capability_name not in capabilities_cache:
                CSVImportService.create_or_get_capability(
                    capability_name, uid_counters, capabilities_cache, subvertical_uid
                )
                stats["capabilities_created"] += 1
            
            cap_uid = capabilities_cache[capability_name]
            
            # 4. Create or get Process
            proc_uid = None
            if process_name:
                process_key = (cap_uid, process_name)
                if process_key not in processes_cache:
                    CSVImportService.create_or_get_process(
                        process_name, process_desc, cap_uid, 
                        uid_counters, processes_cache
                    )
                    stats["processes_created"] += 1
                
                proc_uid = processes_cache[process_key]
                
                # 5. Create or get Subprocess
                subproc_uid = None
                if subprocess_name:
                    subprocess_key = (proc_uid, subprocess_name)
                    if subprocess_key not in subprocesses_cache:
                        CSVImportService.create_or_get_subprocess(
                            subprocess_name, subprocess_desc, proc_uid,
                            uid_counters, subprocesses_cache
                        )
                        stats["subprocesses_created"] += 1
                    
                    subproc_uid = subprocesses_cache[subprocess_key]
                    
                    # 6. Create or get Data Entity
                    de_uid = None
                    if data_entity_name:
                        data_entity_key = (subproc_uid, data_entity_name)
                        if data_entity_key not in data_entities_cache:
                            CSVImportService.create_or_get_data_entity(
                                data_entity_name, data_entity_desc, subproc_uid,
                                uid_counters, data_entities_cache
                            )
                            stats["data_entities_created"] += 1
                        
                        de_uid = data_entities_cache[data_entity_key]
                        
                        # 7. Create or get Data Element
                        if data_element_name:
                            data_element_key = (de_uid, data_element_name)
                            if data_element_key not in data_elements_cache:
                                CSVImportService.create_or_get_data_element(
                                    data_element_name, data_element_desc, de_uid,
                                    uid_counters, data_elements_cache
                                )
                                stats["data_elements_created"] += 1
                    
                    # 8. Create Organization Units and link to Capability
                    if org_units_str:
                        org_units = [ou.strip() for ou in org_units_str.split(',') if ou.strip()]
                        for org_unit_name in org_units:
                            if org_unit_name not in org_units_cache:
                                CSVImportService.create_or_get_org_unit(
                                    org_unit_name, uid_counters, org_units_cache
                                )
                                stats["organization_units_created"] += 1
                            
                            # Link to capability
                            ou_uid = org_units_cache[org_unit_name]
                            CSVImportService.link_capability_to_org_unit(cap_uid, ou_uid)
                    
                    # 9. Create Applications and link to Subprocess
                    if applications_str:
                        applications = [app.strip() for app in applications_str.split(',') if app.strip()]
                        for app_name in applications:
                            if app_name not in applications_cache:
                                CSVImportService.create_or_get_application(
                                    app_name, uid_counters, applications_cache
                                )
                                stats["applications_created"] += 1
                            
                            # Link to subprocess
                            app_uid = applications_cache[app_name]
                            CSVImportService.link_subprocess_to_application(subproc_uid, app_uid)
        
        logger.info(f"CSV import completed: {stats}")
        return stats
        
    except Exception as e:
        logger.error(f"Failed to import CSV to Neo4j: {e}", exc_info=True)
        raise
